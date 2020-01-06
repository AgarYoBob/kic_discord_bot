import discord
import datetime
import openpyxl
import random
import asyncio
import time

client = discord.Client()

@client.event
async def on_ready():
    print("Run_Bot")
    print(client.user.name)
    print(client.user.id)
    print("-------------------------\n")
    file = open('log.txt', 'a', encoding='UTF-8')
    file.write('\n\n==================================================\n')
    file.write('\n[Run Bot]\n')

    #while True:
    #    send_text = input()
    #    await client.send_message(client.get_channel('181663429143035906'), send_text)

    await client.change_presence(game=discord.Game(name='cafe.naver.com/iocommunity', type=1))

    
async def command_log(message):
    now = datetime.datetime.now()
    embed = discord.Embed(title="명령어 사용이 감지되었습니다.", description=message.author.name + " (<@" + message.author.id + ">)\n#" + message.channel.name + " (<#" + message.channel.id + ">)\n**" + message.content + "**", color=0x0000ff)
    embed.set_footer(text=str(now.year) + "년 " + str(now.month) + "월 " + str(now.day) + "일 | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second))
    await client.send_message(client.get_channel('632883852959940608'), embed=embed)
    print('\n   \n%s(%s) | #%s(#%s)\n%s\n%s' % (message.author.name, message.author.id, message.channel, message.channel.id, message.content, str(now.year) + "-" + str(now.month) + "-" + str(now.day) + " | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second)))
    

send_location = 'loc'

@client.event
async def on_message(message):


    if message.channel.id != '181663429143035906' and message.channel.id != '663382190104641551':
        now = datetime.datetime.now()
        file = open('log.txt', 'a', encoding='UTF-8')
        file.write('\n   \n%s(%s) | #%s(#%s)\n%s\n%s' % (message.author.name, message.author.id, message.channel, message.channel.id, message.content, str(now.year) + "-" + str(now.month) + "-" + str(now.day) + " | " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second)))
        
    if message.author.id == '150577293981515776':

        if message.content == '/c list': # /c list
            await client.send_message(message.channel, message.content[3:])

        elif message.content.startswith('/c add '): # /c add
            send_channel = discord.utils.get(server.channels, name = message.content[7:])
            print(send_channel)

            #file = openpyxl.load_workbook("디코_한아커봇_전송채널_목록.xlsx")
            #sheet = file.active
            #member = discord.utils.get(client.get_all_members())
            #for i in range(1, 129):
            #    if str(sheet["B" + str(i)].value) == str(message.author.id):
            #        sheet["A" + str(i)].value = str(message.author.name)
            #        warning_times = str(sheet["C" + str(i)].value)
            #        break
            #    if str(sheet["B" + str(i)].value) == "-":
            #        sheet["A" + str(i)].value = str(message.author.name)
            #        sheet["B" + str(i)].value = str(message.author.id)
            #        sheet["C" + str(i)].value = 0
            #        warning_times = str(sheet["C" + str(i)].value)
            #        break
            #await client.send_message(message.channel, message.content[7:])

        elif message.content.startswith('/c '): # /c set
            file = openpyxl.load_workbook('kic_bot.xlsx')
            sheet = file.active
            sheet['H' + '2'].value = str(message.content[3:])
            await client.send_message(message.channel, '메시지 전송 위치가 <#' + message.content[3:] + '>(으)로 설정되었습니다.')
            file.save('kic_bot.xlsx')

        elif message.content.startswith('/r '):
            #print(send_location)
            file = openpyxl.load_workbook('kic_bot.xlsx')
            sheet = file.active
            await client.send_message(message.channel, '메시지 ' + message.content[3:] + '(을)를 ' + str(sheet["H" + "2"].value) + '에 전송했습니다.')
            await client.send_message(client.get_channel(str(sheet["H" + "2"].value)), message.content[3:])
            file.save('kic_bot.xlsx')
        

client.run(TOKEN)
